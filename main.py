import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
from datetime import datetime
import os
from pathlib import Path
from openpyxl.utils import get_column_letter


def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")

        if key and key not in os.environ:
            os.environ[key] = value


def parse_jira_datetime(value: str | None) -> datetime | None:
    if not value:
        return None

    try:
        return datetime.strptime(value, "%Y-%m-%dT%H:%M:%S.%f%z")
    except ValueError:
        return None


def to_excel_datetime(value: str | None) -> datetime | None:
    parsed = parse_jira_datetime(value)
    if not parsed:
        return None
    return parsed.replace(tzinfo=None)


def extract_sla_metrics(sla_data: dict | None) -> dict:
    if not isinstance(sla_data, dict):
        return {
            "elapsed_minutes": None,
            "elapsed_friendly": None,
            "goal_minutes": None,
            "goal_friendly": None,
            "breached": None,
        }

    cycle = None
    completed_cycles = sla_data.get("completedCycles")
    if isinstance(completed_cycles, list) and completed_cycles:
        cycle = completed_cycles[-1]
    elif isinstance(sla_data.get("ongoingCycle"), dict):
        cycle = sla_data.get("ongoingCycle")

    if not isinstance(cycle, dict):
        return {
            "elapsed_minutes": None,
            "elapsed_friendly": None,
            "goal_minutes": None,
            "goal_friendly": None,
            "breached": None,
        }

    elapsed = cycle.get("elapsedTime") or {}
    goal = cycle.get("goalDuration") or {}

    elapsed_millis = elapsed.get("millis")
    goal_millis = goal.get("millis")

    return {
        "elapsed_minutes": round(elapsed_millis / 60000, 2) if isinstance(elapsed_millis, (int, float)) else None,
        "elapsed_friendly": elapsed.get("friendly"),
        "goal_minutes": round(goal_millis / 60000, 2) if isinstance(goal_millis, (int, float)) else None,
        "goal_friendly": goal.get("friendly"),
        "breached": cycle.get("breached"),
    }


def extract_request_type_details(request_field: dict | None) -> dict:
    if not isinstance(request_field, dict):
        return {
            "request_type_portal_id": None,
            "request_type_portal_name": None,
            "request_type_portal_description": None,
            "service_desk_id": None,
            "portal_id": None,
            "request_current_status": None,
            "request_current_status_category": None,
            "request_current_status_date": None,
        }

    request_type = request_field.get("requestType") or {}
    current_status = request_field.get("currentStatus") or {}
    status_date = current_status.get("statusDate") or {}

    return {
        "request_type_portal_id": request_type.get("id"),
        "request_type_portal_name": request_type.get("name"),
        "request_type_portal_description": request_type.get("description"),
        "service_desk_id": request_type.get("serviceDeskId"),
        "portal_id": request_type.get("portalId"),
        "request_current_status": current_status.get("status"),
        "request_current_status_category": current_status.get("statusCategory"),
        "request_current_status_date": to_excel_datetime(status_date.get("jira")),
    }


def fetch_service_desk_metadata(base_url: str, auth: HTTPBasicAuth, headers: dict, service_desk_id: str) -> dict:
    group_url = f"{base_url}/rest/servicedeskapi/servicedesk/{service_desk_id}/requesttypegroup"
    type_url = f"{base_url}/rest/servicedeskapi/servicedesk/{service_desk_id}/requesttype"

    groups_response = requests.get(group_url, headers=headers, auth=auth)
    groups_response.raise_for_status()
    groups_data = groups_response.json().get("values", [])
    group_map = {str(item.get("id")): item.get("name") for item in groups_data if item.get("id")}

    types_response = requests.get(type_url, headers=headers, auth=auth)
    types_response.raise_for_status()
    types_data = types_response.json().get("values", [])
    type_map = {str(item.get("id")): item for item in types_data if item.get("id")}

    return {
        "groups": group_map,
        "types": type_map,
    }


def calculate_breached_percent(series: pd.Series) -> float | None:
    valid = series.dropna()
    if valid.empty:
        return None
    return round((valid == True).mean() * 100, 2)


def build_group_summary(df: pd.DataFrame, group_column: str, missing_label: str) -> pd.DataFrame:
    temp = df.copy()
    temp[group_column] = temp[group_column].fillna(missing_label)

    grouped = temp.groupby(group_column, as_index=False).agg(
        chamados=("issue_key", "count"),
        tempo_atendimento_horas_media=("tempo_atendimento_horas", "mean"),
        tempo_primeira_resposta_min_media=("tempo_primeira_resposta_min", "mean"),
        sla_resolucao_estourado_pct=("sla_resolucao_estourado", calculate_breached_percent),
        sla_primeira_resposta_estourado_pct=("sla_primeira_resposta_estourado", calculate_breached_percent),
    )

    return grouped.sort_values(by="chamados", ascending=False).round(2)


def format_worksheet(worksheet, date_columns: set[str] | None = None) -> None:
    worksheet.freeze_panes = "A2"

    header_to_index = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=col_idx).value
        if header_value:
            header_to_index[str(header_value)] = col_idx

    if date_columns:
        for column_name in date_columns:
            col_idx = header_to_index.get(column_name)
            if not col_idx:
                continue
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = "dd/mm/yyyy hh:mm"

    for col_idx in range(1, worksheet.max_column + 1):
        max_length = 0
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            text = "" if value is None else str(value)
            if len(text) > max_length:
                max_length = len(text)

        adjusted_width = min(max(max_length + 2, 12), 45)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

# ==============================
# CONFIGURAÇÕES
# ==============================

load_env_file(Path(__file__).with_name(".env"))

JIRA_EMAIL = os.getenv("JIRA_EMAIL")
JIRA_API_TOKEN = os.getenv("JIRA_API_TOKEN")
JIRA_BASE_URL = os.getenv("JIRA_BASE_URL")
PROJECT_KEY = os.getenv("PROJECT_KEY", "TI")

missing_vars = [
    var_name
    for var_name, var_value in {
        "JIRA_EMAIL": JIRA_EMAIL,
        "JIRA_API_TOKEN": JIRA_API_TOKEN,
        "JIRA_BASE_URL": JIRA_BASE_URL,
    }.items()
    if not var_value
]

if missing_vars:
    raise ValueError(
        "Variáveis ausentes no ambiente/.env: " + ", ".join(missing_vars)
    )

# ==============================
# AUTENTICAÇÃO
# ==============================

auth = HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN)

headers = {
    "Accept": "application/json",
    "Content-Type": "application/json"
}

# ==============================
# ENDPOINT DE BUSCA ENHANCED
# ==============================

search_url = f"{JIRA_BASE_URL}/rest/api/3/search/jql"

# ==============================
# JQL — TODOS OS STATUS
# ==============================

jql = f'project = "{PROJECT_KEY}" ORDER BY created ASC'

max_results = 100
all_issues = []
next_page_token = None

print("🔄 Buscando TODOS os chamados do projeto TI...")

# ==============================
# PAGINAÇÃO
# ==============================

while True:
    payload = {
        "jql": jql,
        "maxResults": max_results,
        "fields": [
            "summary",
            "status",
            "issuetype",
            "priority",
            "assignee",
            "reporter",
            "created",
            "updated",
            "resolutiondate",
            "customfield_10010",
            "customfield_10025",
            "customfield_10045",
            "customfield_10046",
            "customfield_10047",
        ]
    }

    if next_page_token:
        payload["nextPageToken"] = next_page_token

    r = requests.post(
        search_url,
        headers=headers,
        auth=auth,
        json=payload
    )

    if r.status_code != 200:
        print("❌ Erro na API Jira")
        print("Status:", r.status_code)
        print("Resposta:", r.text)
        raise Exception("Erro ao buscar chamados")

    data = r.json()
    issues = data.get("issues", [])

    if not issues:
        break

    all_issues.extend(issues)
    next_page_token = data.get("nextPageToken")

    print(f"📄 Chamados coletados: {len(all_issues)}")

    if data.get("isLast", True):
        break

print(f"✅ Total final de chamados: {len(all_issues)}")

service_desk_ids = set()
for issue in all_issues:
    request_data = issue.get("fields", {}).get("customfield_10010")
    if isinstance(request_data, dict):
        request_type = request_data.get("requestType") or {}
        service_desk_id = request_type.get("serviceDeskId")
        if service_desk_id:
            service_desk_ids.add(str(service_desk_id))

service_desk_metadata_cache = {}
for service_desk_id in service_desk_ids:
    try:
        service_desk_metadata_cache[service_desk_id] = fetch_service_desk_metadata(
            JIRA_BASE_URL,
            auth,
            headers,
            service_desk_id,
        )
    except Exception as exc:
        print(f"⚠️ Não foi possível enriquecer metadados do Service Desk {service_desk_id}: {exc}")
        service_desk_metadata_cache[service_desk_id] = {"groups": {}, "types": {}}

# ==============================
# NORMALIZAÇÃO
# ==============================

rows = []

for issue in all_issues:
    f = issue["fields"]

    created_dt = parse_jira_datetime(f.get("created"))
    resolution_dt = parse_jira_datetime(f.get("resolutiondate"))
    first_response_dt = parse_jira_datetime(f.get("customfield_10025"))
    updated_dt = parse_jira_datetime(f.get("updated"))

    tempo_atendimento_horas = None
    tempo_primeira_resposta_min = None

    if created_dt and resolution_dt:
        tempo_atendimento_horas = round((resolution_dt - created_dt).total_seconds() / 3600, 2)

    if created_dt and first_response_dt:
        tempo_primeira_resposta_min = round((first_response_dt - created_dt).total_seconds() / 60, 2)

    sla_resolucao = extract_sla_metrics(f.get("customfield_10045"))
    sla_primeira_resposta = extract_sla_metrics(f.get("customfield_10046"))
    sla_fechamento_apos_resolucao = extract_sla_metrics(f.get("customfield_10047"))
    request_type_details = extract_request_type_details(f.get("customfield_10010"))

    service_desk_id = request_type_details.get("service_desk_id")
    request_type_portal_id = request_type_details.get("request_type_portal_id")

    metadata = service_desk_metadata_cache.get(str(service_desk_id), {"groups": {}, "types": {}})
    group_map = metadata.get("groups", {})
    type_map = metadata.get("types", {})
    request_type_meta = type_map.get(str(request_type_portal_id), {})

    request_type_group_ids = request_type_meta.get("groupIds") if isinstance(request_type_meta, dict) else None
    if isinstance(request_type_group_ids, list):
        request_type_group_names = [group_map.get(str(group_id), str(group_id)) for group_id in request_type_group_ids]
    else:
        request_type_group_names = []

    request_type_group_name = request_type_group_names[0] if request_type_group_names else None
    request_type_full_path = None
    if request_type_details.get("request_type_portal_name"):
        if request_type_group_name:
            request_type_full_path = f"{PROJECT_KEY}/{request_type_group_name}/{request_type_details.get('request_type_portal_name')}"
        else:
            request_type_full_path = f"{PROJECT_KEY}/{request_type_details.get('request_type_portal_name')}"

    rows.append({
        "issue_key": issue["key"],
        "summary": f.get("summary"),
        "request_type_id": f.get("issuetype", {}).get("id"),
        "status": f.get("status", {}).get("name"),
        "status_category": f.get("status", {}).get("statusCategory", {}).get("name"),
        "request_type": f.get("issuetype", {}).get("name"),
        "request_type_description": f.get("issuetype", {}).get("description"),
        "request_type_subtask": f.get("issuetype", {}).get("subtask"),
        "request_type_hierarchy_level": f.get("issuetype", {}).get("hierarchyLevel"),
        "request_type_group_name": request_type_group_name,
        "request_type_group_names": ", ".join(request_type_group_names) if request_type_group_names else None,
        "request_type_full_path": request_type_full_path,
        "priority": f.get("priority", {}).get("name") if f.get("priority") else None,
        "assignee": f.get("assignee", {}).get("displayName") if f.get("assignee") else None,
        "reporter": f.get("reporter", {}).get("displayName") if f.get("reporter") else None,
        "created": created_dt.replace(tzinfo=None) if created_dt else None,
        "updated": updated_dt.replace(tzinfo=None) if updated_dt else None,
        "resolution_date": resolution_dt.replace(tzinfo=None) if resolution_dt else None,
        "first_response_date": first_response_dt.replace(tzinfo=None) if first_response_dt else None,
        "tempo_atendimento_horas": tempo_atendimento_horas,
        "tempo_primeira_resposta_min": tempo_primeira_resposta_min,
        **request_type_details,
        "sla_resolucao_min": sla_resolucao["elapsed_minutes"],
        "sla_resolucao_texto": sla_resolucao["elapsed_friendly"],
        "sla_resolucao_meta_min": sla_resolucao["goal_minutes"],
        "sla_resolucao_meta_texto": sla_resolucao["goal_friendly"],
        "sla_resolucao_estourado": sla_resolucao["breached"],
        "sla_primeira_resposta_min": sla_primeira_resposta["elapsed_minutes"],
        "sla_primeira_resposta_texto": sla_primeira_resposta["elapsed_friendly"],
        "sla_primeira_resposta_meta_min": sla_primeira_resposta["goal_minutes"],
        "sla_primeira_resposta_meta_texto": sla_primeira_resposta["goal_friendly"],
        "sla_primeira_resposta_estourado": sla_primeira_resposta["breached"],
        "sla_fechamento_apos_resolucao_min": sla_fechamento_apos_resolucao["elapsed_minutes"],
        "sla_fechamento_apos_resolucao_texto": sla_fechamento_apos_resolucao["elapsed_friendly"],
        "sla_fechamento_apos_resolucao_meta_min": sla_fechamento_apos_resolucao["goal_minutes"],
        "sla_fechamento_apos_resolucao_meta_texto": sla_fechamento_apos_resolucao["goal_friendly"],
        "sla_fechamento_apos_resolucao_estourado": sla_fechamento_apos_resolucao["breached"],
    })

df = pd.DataFrame(rows)

# ==============================
# RESUMOS
# ==============================

resumo_geral = pd.DataFrame([
    {
        "total_chamados": len(df),
        "chamados_resolvidos": int(df["resolution_date"].notna().sum()),
        "tempo_atendimento_horas_media": round(df["tempo_atendimento_horas"].mean(), 2) if df["tempo_atendimento_horas"].notna().any() else None,
        "tempo_atendimento_horas_mediana": round(df["tempo_atendimento_horas"].median(), 2) if df["tempo_atendimento_horas"].notna().any() else None,
        "tempo_primeira_resposta_min_media": round(df["tempo_primeira_resposta_min"].mean(), 2) if df["tempo_primeira_resposta_min"].notna().any() else None,
        "tempo_primeira_resposta_min_mediana": round(df["tempo_primeira_resposta_min"].median(), 2) if df["tempo_primeira_resposta_min"].notna().any() else None,
        "sla_resolucao_estourado_pct": calculate_breached_percent(df["sla_resolucao_estourado"]),
        "sla_primeira_resposta_estourado_pct": calculate_breached_percent(df["sla_primeira_resposta_estourado"]),
    }
])

resumo_por_responsavel = build_group_summary(df, "assignee", "Sem responsável")
resumo_por_status = build_group_summary(df, "status", "Sem status")
resumo_por_tipo_completo = build_group_summary(df, "request_type_full_path", "Sem tipo completo")

# ==============================
# EXPORTAÇÃO
# ==============================

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
arquivo = f"jira_TI_todos_chamados_{timestamp}.xlsx"

with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="dados", index=False)
    resumo_geral.to_excel(writer, sheet_name="resumo_geral", index=False)
    resumo_por_responsavel.to_excel(writer, sheet_name="resumo_responsavel", index=False)
    resumo_por_status.to_excel(writer, sheet_name="resumo_status", index=False)
    resumo_por_tipo_completo.to_excel(writer, sheet_name="resumo_tipo_completo", index=False)

    format_worksheet(
        writer.sheets["dados"],
        date_columns={
            "created",
            "updated",
            "resolution_date",
            "first_response_date",
            "request_current_status_date",
        },
    )
    format_worksheet(writer.sheets["resumo_geral"])
    format_worksheet(writer.sheets["resumo_responsavel"])
    format_worksheet(writer.sheets["resumo_status"])
    format_worksheet(writer.sheets["resumo_tipo_completo"])

print(f"📁 Planilha gerada com sucesso: {arquivo}")